"""
Drugi dio: Logistička Regresija.
Prvo je potrebno vidjeti kako se kreće točnost modela na različitim podacima te pronaći one koji pokazuju
najbolju točnost modela. Paralelno s time potrebno je za svake podatke naći najbolje odgovarajući
hiperparametar C=1/lambda.
C - koeficijent dopuštene pogreške klasifikacije - obrnuto proporcionalni jačini regularizacije lambda
Potrebno je ispitati kombinacije različith C i setova podataka
"""

from Preprocessing import input_data, output_data
# Ukoliko se žele stvarati novi podaci svaku skriptu, ovo NE treba biti zakomentirano
from Preprocessing import divided_train_data, all_X_test_data
from Preprocessing import Y_train


import openpyxl, pickle, math, os, shutil
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score


output_Log_Reg = output_data+"B_Logistic_Regression/"       # folder gdje se spremaju rezultati
xlsx_file = output_Log_Reg+"Acc_LogReg.xlsx"                # ime excel file u koji se spremaju rezultati

# Ako ne postoji folder stvara novi, ako postoji briše datoteke
with open(input_data+'/divided_train_data.pickle', 'rb') as f_X_train:
    divided_train_data = pickle.load(f_X_train)
with open(input_data+'/all_X_test_data.pickle', 'rb') as f_test:
    all_X_test_data = pickle.load(f_test)
try:
    shutil.rmtree(output_Log_Reg)
except:
    FileNotFoundError
os.mkdir(output_Log_Reg)


# Funkcija koja računa točnost modela za određene podatke i hiperparametar C
def logreg_f(C, X_train, Y_train, X_valid, Y_valid):
    logreg_model = LogisticRegression(C=C, max_iter=1e7)        # Broj iteracija povećan zbog poly podataka
    logreg_model.fit(X_train, Y_train)
    acc_train = round(logreg_model.score(X_train, Y_train) * 100, 2)
    prediction = logreg_model.predict(X_valid)
    acc_valid = round(accuracy_score(prediction, Y_valid) *100,2)
    # error = round(sklearn.metrics.mean_squared_error(Y_valid, prediction), 2)  # kvadratna razlika 2 vektora
    return [acc_train, acc_valid]


C_range = [10**i for i in range(-10,5)]      # raspon C hiperparametara
# C_range = [10**i for i in range(-10,0)]      #PROBA

accuracy_dict = {}                           # spremanje parova (C:točnost za svaki C) za cijeli set podataka
best_grid_acc = {}                           # spremanje najbolje točnosti za podatke

# Iteriranje spremljenih podataka i analiza
for data_name in divided_train_data["X_train_data"]:
    X_train = divided_train_data["X_train_data"][data_name]
    X_valid = divided_train_data["X_valid_data"][data_name]

    Y_train = divided_train_data["Y_train_data"][data_name]
    Y_valid = divided_train_data["Y_valid_data"][data_name]

    max_acc = 0
    best_grid_acc[data_name], acc_for_data = [], []

    for C in C_range:
        accurities = logreg_f(C, X_train, Y_train, X_valid, Y_valid)
        acc_for_data.append(accurities)

        if accurities[1] > max_acc:                   # spremanje nove maksimalne točnosti validacijskog seta
            n_C = C_range.index(C)
            max_acc = accurities[1]
            best_grid_acc[data_name] = [n_C]
        elif accurities[1] == max_acc:                # ako točnost već postoji, za nju se dodaju pripadajući hiperparametar C
            n_C = C_range.index(C)
            best_grid_acc[data_name].append(n_C)

    accuracy_dict[data_name] = acc_for_data         # dodavanje svih točnosti u rječnik

accuracy_DF = pd.DataFrame(accuracy_dict, index=[C_range])  # kreiranje DataFrame-project_data





# Funkcija koja zapisuje sve točnosti u excel
def to_excel():
    accuracy_DF.to_excel(xlsx_file, sheet_name="Accuracity")                     # zapisivanje excel file-project_data
    # Zapisivanje oznake za C u tablicu i bojanje oznake
    workbook = openpyxl.load_workbook(xlsx_file)
    workbook.get_sheet_by_name("Accuracity").cell(row=1, column=1).value = "C"
    workbook.get_sheet_by_name("Accuracity").cell(row=1, column=1).fill = \
        openpyxl.styles.PatternFill("solid", fgColor="00FFFF00")  # bojanje oznaka gamma, C

    # Bojanje ćelija u excelu tamo gdje je najveća točnost modela za pojedine podatke
    for data_name in accuracy_DF:
        n_col = accuracy_DF.columns.get_loc(data_name)          # idex stupca

        for n_row in (best_grid_acc[data_name]):                # idex retka
            workbook.get_sheet_by_name("Accuracity").cell(row=n_row+2, column=n_col+2).fill = \
                openpyxl.styles.PatternFill("solid", fgColor="00FF0000")
    workbook.save(xlsx_file)

to_excel()



"""
Funkcija plota točnost setova za treniranje i validaciju, za standardne podatke i poly4 podatke (ručno odabrano).
Svi skalirani i PCA podaci imaju istu tendenciju rasta kao i standardni dok poly3 i poly 4 drugačije pa su iz tog 
razloga odabrani za graf.
"""
def plot_accuracity(data_sets):
    C_log = [math.log(i[0], 10) for i in accuracy_DF.index]     # logaritam(C) s bazom 10
    for data in data_sets:

        valid_acc = list(zip(*accuracy_DF[data]))[1]        # vadi prvi element iz liste lista
        train_acc = list(zip(*accuracy_DF[data]))[0]

        color = next(plt.gca()._get_lines.prop_cycler)['color']
        plt.plot(C_log, valid_acc,linestyle='-', c=color, label=data)
        plt.plot(C_log, train_acc, linestyle=':',  c=color)
        plt.grid(color='k', linestyle=':', linewidth=0.5)

        fig = plt.gcf()
        plt.xlabel("log(C) [-]")
        plt.ylabel("Accuracity [%]")
        plt.title("Validation and Train acc comparison")
        plt.draw()
    plt.legend()
    fig.savefig(output_Log_Reg+"plot_LogReg.png", dpi=300)

# plot_accuracity(data_sets= ["X", "poly3_X", "poly4_X"])
plot_accuracity(data_sets= ["X", "scal_std_X", "scal_MM_X"])



# Funkcija kojia ispisuje koeficijente korelacije (važnost svake kategorije u modelu)
def Log_Reg_coeffs():
    X_train = divided_train_data["X_train_data"]["X"]
    X_valid = divided_train_data["X_valid_data"]["X"]
    Y_train = divided_train_data["Y_train_data"]["X"]
    Y_valid = divided_train_data["Y_valid_data"]["X"]

    logreg_model = LogisticRegression(C=C, max_iter=1e7)        # Broj iteracija povećan zbog poly podataka
    logreg_model.fit(X_train, Y_train)
    acc_train = round(logreg_model.score(X_train, Y_train) * 100, 2)
    prediction = logreg_model.predict(X_valid)
    acc_valid = round(accuracy_score(prediction, Y_valid) *100,2)

    coeff_df = pd.DataFrame(X_train.columns)
    coeff_df.columns = ['Feature']
    coeff_df["Correlation"] = pd.Series(logreg_model.coef_[0])
    coeff_df = coeff_df.sort_values(by='Correlation', key=abs, ascending=False)
    coeff_df = coeff_df.reset_index(drop=True)
    print(coeff_df)

# Log_Reg_coeffs()






"""
Zaključci:
    - Nakon ispitivanja zaključak je da su za različite podatke različiti hiperparametri C najbolji izbor,
        ovisno o podacima.

    - Točnost modela dosta ovisi o odabiru random_state-project_data i podacima koji se uzimaju za train i validation set!

    - Najbolju točnost train seta pokazuju podaci koji su preslikani u višu dimenziju i to za 4 razine: točnost 88.23%
        -> logično jer je model s više parametara najlakše postaje prenaučen. 

    - Dobivaju se značajne razlike u točnosti za različite podatke od seta do seta podataka. 
        Može li se to kako popraviti?

    - Treba imati na umu da su ovdje svi podaci prebačeni u kategoričke varijable
            
    - Nakon provedene PCA analize dobiva se nešto manja točnost modela koja se manje smanjuje sa smanjivanjem
        broja dimenzija. Međutim konvergencija je značajno brža: koristan alat za veće skupove podataka 
   
    -Nema smisla dizati podatke u viši prostor značjki i potom ih skalirati 
    
    - Za pravi set podataka može se očekivati malo manja točnost jer na dosta podataka fali kategorija: Age
       
    - Odabrani podaci: prvotni X set podaci i C=100 
    
"""














