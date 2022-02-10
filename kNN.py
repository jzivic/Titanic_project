"""
Četvrti dio: kNN algoritam.
Jedina nepoznanica je pronaći optimalan broj najbližih susjeda na različitim ulaznim podacima koji će se varirati
za sve setove podataka u rangu 1-10.
"""


from Preprocessing import input_data, output_data
from Preprocessing import Y_train
from Preprocessing import divided_train_data, all_X_test_data
from sklearn.metrics import accuracy_score

import openpyxl, pickle, os, shutil, math, sklearn
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.neighbors import KNeighborsClassifier


output_D = output_data+"D_kNN/"

with open(input_data+'/divided_train_data.pickle', 'rb') as f_X_train:
    divided_train_data = pickle.load(f_X_train)
with open(input_data+'/all_X_test_data.pickle', 'rb') as f_test:
    all_X_test_data = pickle.load(f_test)

try:
    shutil.rmtree(output_D)
except:
    FileNotFoundError
os.mkdir(output_D)



# Funkcija koja računa točnost ovisno za odabrane podatke i broj susjeda
def kNN_f(n_neighbors, X_train, Y_train, X_valid, Y_valid):
    knn_model = KNeighborsClassifier(n_neighbors=n_neighbors)
    knn_model.fit(X_train, Y_train)
    acc_train = round(knn_model.score(X_train, Y_train) * 100, 2)
    prediction = knn_model.predict(X_valid)
    acc_valid = round(accuracy_score(prediction, Y_valid) *100,2)

    return [acc_train, acc_valid]


xlsx_file = output_D+"Acc_kNN.xlsx"  # ime excel file u koji se spremaju rezultati, kasnije
n_neigh_range = [i for i in range(1,41)]    # raspon za koji se računa analiza kNN broj susjeda
accuracy_dict, best_n_neighb_acc = {}, {}    # parovi (n_neigh:točnost za svaki) te najbolje točnosti za data set

# iteracija kNN analiza
for data_name in divided_train_data["X_train_data"]:
    X_train = divided_train_data["X_train_data"][data_name]
    Y_train = divided_train_data["Y_train_data"][data_name]
    X_valid = divided_train_data["X_valid_data"][data_name]
    Y_valid = divided_train_data["Y_valid_data"][data_name]

    max_acc = 0
    best_n_neighb_acc[data_name], acc_for_data = [], []
    for n_neighb in n_neigh_range:
        accurities = kNN_f(n_neighb, X_train, Y_train, X_valid, Y_valid)
        acc_for_data.append(accurities)

        if accurities[1] > max_acc:                   # spremanje nove maksimalne točnosti validacijskog seta
            max_acc = accurities[1]
            best_n_neighb_acc[data_name] = [n_neighb]

        elif accurities[1] == max_acc:                # ako točnost već postoji, za nju se dodaju pripadajući hiperparametar C
            best_n_neighb_acc[data_name].append(n_neighb)

    accuracy_dict[data_name] = acc_for_data  # dodavanje svih točnosti u rječnik


accuracy_DF = pd.DataFrame(accuracy_dict, index=[n_neigh_range])  # kreiranje DataFrame-a
accuracy_DF.to_excel(xlsx_file, sheet_name="Accuracity")                     # zapisivanje excel file-a




# Zapisivanje oznake za n neighb u tablicu i bojanje oznake
def to_ecxel():
    workbook = openpyxl.load_workbook(xlsx_file)
    workbook.get_sheet_by_name("Accuracity").cell(row=1, column=1).value = "n Neigh"
    workbook.get_sheet_by_name("Accuracity").cell(row=1, column=1).fill = \
        openpyxl.styles.PatternFill("solid", fgColor="00FFFF00")  # bojanje oznaka gamma, C

    # traženje i bojanje maksimalne točnosti za svaki data set
    for data_name in accuracy_DF:
        n_col = accuracy_DF.columns.get_loc(data_name)          # idex stupca

        for n_row in (best_n_neighb_acc[data_name]):                # idex retka
            workbook.get_sheet_by_name("Accuracity").cell(row=n_row+1, column=n_col+2).fill = \
                openpyxl.styles.PatternFill("solid", fgColor="00FF0000")
    workbook.save(xlsx_file)
to_ecxel()





def plot_accuracity(data_sets):
    for data in data_sets:
        valid_acc = list(zip(*accuracy_DF[data]))[1]
        train_acc = list(zip(*accuracy_DF[data]))[0]
        color = next(plt.gca()._get_lines.prop_cycler)['color']
        plt.plot(n_neigh_range, valid_acc,linestyle='-', c=color, label=data)
        plt.plot(n_neigh_range, train_acc, linestyle=':',  c=color)
        plt.grid(color='k', linestyle=':', linewidth=0.5)

        fig = plt.gcf()
        plt.xlim(0, max(n_neigh_range)+1)
        plt.xlabel("n_neighbors [-]")
        plt.ylabel("Accuracity [%]")
        plt.title("Validation and train accuracity comparison")
        plt.draw()

    plt.legend()
    fig.savefig(output_D+"plot_kNN.png", dpi=300)


plot_accuracity(data_sets= [data for data in divided_train_data["X_train_data"]])
# plot_accuracity(data_sets= ["X", "poly4_X"])
# plot_accuracity(data_sets= ["X", "scal_std_X"])








"""
Zaključak:
    -D_kNN model je dosta osjetljiv na promjene podataka unutar split metode!
   
    -Na primjeru neskaliranih podataka kao i većine ostalih zaključuje se da najbolju točnost za train set imaju modeli
     u kojima se uzima samo 1 najbliži susjed (osim podataka s Min Max skaliranjem)
    
    -Najveću točnost train seta očekivano ima model s podacima dizanima u viši prostor značajki: preko 85%. 

    -Zanimljivo je da se na nekim primjerima mijenja kontinuirano / proporcionalno s brojem susjeda (preskače)
    
    - Broj parametara će se uzeti 20-25 na temelju pregleda validacijskog seta jer se pokazao kao najstabilniji
    
    -Najbolju točnost imaju MM podaci za 15-20 broj susjeda
    
"""











