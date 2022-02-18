"""
Treći dio: Support Vector Machine
Potrebno je provesti pretraživanje po rešetki nad hiperparametrima C i gamma kako bi se dobile optimalne vrijednosti
za sve setove podataka. Očekuje se da će oni biti malo drugačiji za različite setove.
C - dozvoljena pogreška klasifikacije; obrnuto proporcionalna jačina regularizacije lambda, C=1/lambda
gamma - koef. jezgre za rbf, poly i sigmoidalnu funkciju.
Analiza će se provjeriti posebno za 3 jezgrene funkcije: rbf, poly i linear.
Na kraju će se rezultati zapisati u excel datoteke posebno za svaki set podataka, u svakoj za sve 3 jezgrene funkcije.

"""


from Preprocessing import input_data, output_data
# Ukoliko se žele stvarati novi podaci svaku skriptu, ovo NE treba biti zakomentirano
# fromPreprocessing import Y_train
# fromPreprocessing import divided_train_data, all_X_test_data

import openpyxl, pickle, os, shutil
import pandas as pd
from sklearn.svm import SVC
from sklearn.metrics import accuracy_score

output_SVM = output_data+"C_SVM/"

with open(input_data+'/divided_train_data.pickle', 'rb') as f_X_train:
    divided_train_data = pickle.load(f_X_train)
with open(input_data+'/all_X_test_data.pickle', 'rb') as f_test:
    all_X_test_data = pickle.load(f_test)

try:
    shutil.rmtree(output_SVM)
except:
    FileNotFoundError
os.mkdir(output_SVM)


# Velika SVM funkcija koja računa točnost za rbf, poly i linear jezgrene funkcije
def grid_search(data_name, X_train, Y_train, X_valid, Y_valid):

    # Pomoćna funkcija koja ispisuje koliko je gotovo simulacija od ukupnog broja
    def current_sim_number(c,gamma, C_range, gamma_range):
        m = C_range.index(c)
        n = gamma_range.index(gamma)        # indexi trenutnog položaja

        num = n * len(C_range) + m +1       # number of simulation
        percentage_over = num / (len(C_range) * len(gamma_range))
        percentage_over_str = (str(num) + "/" + str(len(C_range) * len(gamma_range)))
        print(percentage_over_str)
        return percentage_over


    # Postavljanje hiperparametara i excel datoteke za spremanje pretraživanja
    xlsx_name = output_SVM+"Acc_SVM_"+data_name+".xlsx"                    # ime excel file u koji se spremaju rezultati, kasnije
    Acc_grid_search = pd.ExcelWriter(xlsx_name)             # stvaranje excela
    C_range = range(-10, 1)                                 # rang baze za C
    gamma_range = range(-20, 1)
    C_f = lambda m: 2 ** m                                  # Funkcije za računanje C i gamma
    gamma_f = lambda n: 2 ** n
    best_grid_acc = {"radial_base":[], "poly":[], "linear":[]}

    # isprobavanje jel radi
    # C_range = range(-5, -1)  # Pretraživanje se provodi u ovim rangovima     # proba
    # gamma_range = range(1, 5)
    # best_grid_acc = {"radial_base":[]}


    def radial_base_function():
        accuracity_matrix = {}
        max_acc = 0                           # inicijalno postavljanje najbolje točnosti i array-project_data pozicija

        for gamma_ in gamma_range:                          # Pretraživanje za jedan red (C=..)
            gamma_str = "{:.2e}".format(gamma_f(gamma_))
            accuracity_matrix[gamma_str] = []               # Rezultati se spremaju u matricu

            for c_ in C_range:                          # Pretraživanje za jedan stupac (gamma=..)
                svcModel = SVC(C=C_f(c_), kernel="rbf", gamma=gamma_f(gamma_))
                svcModel.fit(X_train, Y_train)
                acc_train = round(svcModel.score(X_train, Y_train) * 100, 2)
                prediction = svcModel.predict(X_valid)
                acc_valid = round(accuracy_score(prediction, Y_valid) * 100, 2)

                if acc_valid > max_acc:                                            # traženje najbolje točnosti
                    n_gamma, n_C = gamma_range.index(gamma_), C_range.index(c_)
                    max_acc = acc_valid
                    best_grid_acc["radial_base"] = [[n_C, n_gamma]]
                elif acc_valid == max_acc:                                      # nadopunjavanje array-project_data za naj točnost
                    n_gamma, n_C = gamma_range.index(gamma_), C_range.index(c_)
                    best_grid_acc["radial_base"].append([n_C, n_gamma])

                accuracity_matrix[gamma_str].append([acc_train, acc_valid])                # punjenje stupaca DataFrame-project_data
                # current_sim_number(c_, gamma_, C_range, gamma_range)       # ispisuje gotove simulacije

        ind_names = ["{:.2e}".format(C_f(i)) for i in C_range]          # imena indexa DataFrame-project_data, gamma=.. podaci
        accuracity_matrix = pd.DataFrame(accuracity_matrix, index=ind_names)

        # Redovi su vrijednosti C, kolone vrijednosti gamma
        accuracity_matrix.to_excel(Acc_grid_search, sheet_name="radial_base", startcol=0, startrow=0)

    radial_base_function()




    def poly_function():
        C_range_poly = range(-10, 3)
        gamma_range_poly = range(-10, -1)
        # gamma_range_poly = range(-10, -9)       # pomoćne vrijednosti za isprobavanje
        # C_range_poly = range(-10, -9)

        accuracity_matrix = {}
        max_acc = 0
        for gamma_ in gamma_range_poly:
            gamma_str = "{:.2e}".format(gamma_f(gamma_))
            accuracity_matrix[gamma_str] = []

            for c_ in C_range_poly:
                svcModel = SVC(C=C_f(c_), kernel="poly", gamma=gamma_f(gamma_))
                svcModel.fit(X_train, Y_train)
                acc_train = round(svcModel.score(X_train, Y_train) * 100, 2)
                prediction = svcModel.predict(X_valid)
                acc_valid = round(accuracy_score(prediction, Y_valid) * 100, 2)
                # current_sim_number(c_, gamma_, C_range_poly, gamma_range_poly)

                if acc_valid > max_acc:                                            # traženje najbolje točnosti
                    n_gamma, n_C = gamma_range_poly.index(gamma_), C_range_poly.index(c_)
                    max_acc = acc_valid
                    best_grid_acc["poly"] = [[n_C, n_gamma]]

                elif acc_valid == max_acc:                                      # nadopunjavanje array-project_data za naj točnost
                    n_gamma, n_C = gamma_range_poly.index(gamma_), C_range_poly.index(c_)
                    best_grid_acc["poly"].append([n_C, n_gamma])

                accuracity_matrix[gamma_str].append([acc_train, acc_valid])  # punjenje stupaca DataFrame-project_data

        ind_names = ["{:.2e}".format(C_f(i)) for i in C_range_poly]
        accuracity_matrix = pd.DataFrame(accuracity_matrix, index=ind_names)
        accuracity_matrix.to_excel(Acc_grid_search, sheet_name="poly", startcol=0, startrow=0)

    poly_function()




    def linear_function():
        gamma_range_linear = [1]
        C_range_linear = range(-15, 2)
        # C_range_linear = range(-15, -10)

        accuracity_matrix = {}
        max_acc = 0
        for gamma_ in gamma_range_linear:
            accuracity_matrix[1] = []

            for c_ in C_range_linear:
                svcModel = SVC(C=C_f(c_), kernel="linear")
                svcModel.fit(X_train, Y_train)
                acc_train = round(svcModel.score(X_train, Y_train) * 100, 2)
                prediction = svcModel.predict(X_valid)
                acc_valid = round(accuracy_score(prediction, Y_valid) * 100, 2)
                accuracity_matrix[1].append([acc_train, acc_valid])
                # current_sim_number(c_, gamma_, C_range_linear, gamma_range_linear)

                if acc_valid > max_acc:                                            # traženje najbolje točnosti
                    n_gamma, n_C = gamma_range_linear.index(gamma_), C_range_linear.index(c_)
                    max_acc = acc_valid
                    best_grid_acc["linear"] = [[n_C, n_gamma]]
                elif acc_valid == max_acc:                                      # nadopunjavanje array-project_data za naj točnost
                    n_gamma, n_C = gamma_range_linear.index(gamma_), C_range_linear.index(c_)
                    best_grid_acc["linear"].append([n_C, n_gamma])

        ind_names = ["{:.2e}".format(C_f(i)) for i in C_range_linear]
        accuracity_matrix = pd.DataFrame(accuracity_matrix, index=ind_names)
        accuracity_matrix.to_excel(Acc_grid_search, sheet_name="linear", startcol=0, startrow=0)        # Redovi su vrijednosti C, kolone vrijednosti gamma

    linear_function()


    Acc_grid_search.save()                              # Spremanje DF u excel file
    workbook=openpyxl.load_workbook(xlsx_name)

    # Otvara sve sheetove ovisno o rječniku gdje su spremljeni podaci
    for kernel_function in best_grid_acc.keys():
        workbook.get_sheet_by_name(kernel_function).cell(row=1, column=1).value = "C/gamma"
        workbook.get_sheet_by_name(kernel_function).cell(row=1, column=1).fill = \
            openpyxl.styles.PatternFill("solid", fgColor="00FFFF00")                   # bojanje oznaka gamma, C

        # iteracija po zapisanim koordinatama gdje se nalazi maksimalna točnost modela
        for position in best_grid_acc[kernel_function]:
            workbook.get_sheet_by_name(kernel_function).cell(row=2+position[0], column=2+position[1]).fill =\
                openpyxl.styles.PatternFill("solid", fgColor="00FF0000")

    workbook.save(xlsx_name)


####################################################################################################################



# for item in all_X_train_data.items():        # hvata sve podatke iz rječnika

for data_name in divided_train_data["X_train_data"]:

    X_train = divided_train_data["X_train_data"][data_name]
    Y_train = divided_train_data["Y_train_data"][data_name]
    X_valid = divided_train_data["X_valid_data"][data_name]
    Y_valid = divided_train_data["Y_valid_data"][data_name]

    grid_search(data_name, X_train, Y_train, X_valid, Y_valid)          # provodi cijelu analizu i izbacuje exelc




####################################################################################################################






"""
Zaključak:
    -Točnost osnovnog inicijalnog modela train seta SVC je 81.5% project_data modela s linearnom jezgrenom funkcijom 77%. 

    -Točnost modela s rbf jezgrenom funkcijom je preko 88.23%. Negativna stvar je da za veće 
        vrijednosti C i gamma hiperparametara model sporije konvergira

    -Nakon provedene analize zaključuje se da su veće vrijednosti hiperparametara optimalne. 
    -C=1, gamma=1 su vrijednosti hiperparametara za koje model s osnovnim pokacima ima najbolju točnost
    
    -Rezultati pravilno konvergiraju rješenju kada se vrijednosti hiperparametara mijenjaju po kvadratnoj funkciji 
       
    -Modeli s drugim podacima i funkcijama imaju istu maksimalnu točnost, ALI se do nje može brže doći promjenom HP.
        -> pogotovo kod korištenja poly jezgrene funkcije
        
    -Vjerojatno se točnost kod poly modela može još malo podići uz dosta veću računalnu zahtjevnost - neisplativo




    # za čiste podatke odabrano:
     0.5 / 0.125 - teško točno pogoditi jer se mijenjaju od seta do seta podataka

    # poly4: stabilniji rezultati
    0.25/3e-6


"""









# Funkcija koja ispisuje tablicu s koeficijentima korelacije (vaćnost svake kategorije za predviđanje)
def SVC_coeffs():

    X_train = divided_train_data["X_train_data"]["X"]
    Y_train = divided_train_data["Y_train_data"]["X"]
    X_valid = divided_train_data["X_valid_data"]["X"]
    Y_valid = divided_train_data["Y_valid_data"]["X"]

    # svcModel = SVC(C=0.5, kernel="rbf", gamma=0.125)
    svcModel = SVC(C=33, kernel="linear")
    svcModel.fit(X_train, Y_train)
    # acc_train = round(svcModel.score(X_train, Y_train) * 100, 2)
    # prediction = svcModel.predict(X_valid)
    # acc_valid = round(accuracy_score(prediction, Y_valid) * 100, 2)



    coeff_df = pd.DataFrame(X_train.columns)
    coeff_df.columns = ['Feature']
    coeff_df["Correlation"] = pd.Series(svcModel.coef_[0])
    coeff_df = coeff_df.sort_values(by='Correlation', key=abs, ascending=False)
    coeff_df = coeff_df.reset_index(drop=True)

    print(coeff_df)

    return coeff_df


SVC_coeffs()


    # coeff_df = pd.DataFrame(train_X.columns)
    # coeff_df.columns = ['Feature']
    # coeff_df["Correlation"] = pd.Series(svc.coef_[0])
    # coeff_df = coeff_df.sort_values(by='Correlation', key=abs, ascending=False)
    # coeff_df = coeff_df.reset_index(drop=True)